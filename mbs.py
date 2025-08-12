import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import logging
import os
import glob

from tvr_processor import extract_tvr_data  # ✅ IMPORT the TVR extractor

logger = logging.getLogger(__name__)

def safe_get_cell(df, row, col, default=0):
    try:
        value = df.iloc[row, col]
        return value if not pd.isna(value) else default
    except IndexError:
        logger.warning(f"Warning: Index [{row},{col}] out of bounds for dataframe with shape {df.shape}")
        return default

def safe_set_cell(ws, cell_ref, value):
    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            cell_ref = merged_range.start_cell.coordinate
            break
    ws[cell_ref] = value
    return cell_ref

def process_excel_data(input_a_path, input_b_path, skeleton_path, output_path):
    logger.info(f"Process started on {datetime.now().strftime('%A, %B %d, %Y at %H:%M:%S')}")

    # File checks
    for path, name in [(input_a_path, "Non Cricket Input"), (input_b_path, "TVR Output"), (skeleton_path, "Skeleton")]:
        if not os.path.exists(path):
            error_msg = f"{name} file not found at path: {path}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)


    try:
    # Load data
        property_details = pd.read_excel(input_a_path, sheet_name="Property Details", header=None)
        channel_platform = pd.read_excel(input_a_path, sheet_name="Channel & Platform Details", header=None)
        program_performance = pd.read_excel(input_a_path, sheet_name="Program Performance", header=None)
        input_b = pd.read_excel(input_b_path, header=None)
    except Exception as e:
        logger.error(f"Error loading input files: {str(e)}")
        raise
    
    wb = load_workbook(skeleton_path)
    sheet1 = wb[wb.sheetnames[0]]
    sheet2 = wb[wb.sheetnames[1]]

    # Same extraction logic as before ...
    prop_b1 = safe_get_cell(property_details, 0, 1, "")
    prop_b8 = safe_get_cell(property_details, 7, 1, "")
    prop_b14 = safe_get_cell(property_details, 13, 1, "")
    prop_b29 = safe_get_cell(property_details, 28, 1, "")
    prop_a29 = safe_get_cell(property_details, 28, 0, "")
    prop_b3 = safe_get_cell(property_details, 2, 1, "")
    prop_b4 = safe_get_cell(property_details, 3, 1, "")
    prop_b7 = safe_get_cell(property_details, 6, 1, "")
    prop_b9 = safe_get_cell(property_details, 8, 1, "")
    prop_b10 = safe_get_cell(property_details, 9, 1, "")
    prop_b11 = safe_get_cell(property_details, 10, 1, "")
    prop_b12 = safe_get_cell(property_details, 11, 1, 2)
    prop_b13 = safe_get_cell(property_details, 12, 1, 1)
    prop_b20 = safe_get_cell(property_details, 19, 1, "")
    prop_b21 = safe_get_cell(property_details, 20, 1, "")
    prop_b22 = safe_get_cell(property_details, 21, 1, "")
    prop_b23 = safe_get_cell(property_details, 22, 1, "")
    prop_b26 = safe_get_cell(property_details, 25, 1, "")
    prop_b27 = safe_get_cell(property_details, 26, 1, "")
    prop_b28 = safe_get_cell(property_details, 27, 1, "")
    prop_b32 = safe_get_cell(property_details, 31, 1, 0)

    channel_b5 = safe_get_cell(channel_platform, 4, 1, "")
    channel_c5 = safe_get_cell(channel_platform, 4, 2, "")
    channel_c6 = safe_get_cell(channel_platform, 5, 2, "")
    channel_c7 = safe_get_cell(channel_platform, 6, 2, "")
    channel_c8 = safe_get_cell(channel_platform, 7, 2, "")
    channel_c9 = safe_get_cell(channel_platform, 8, 2, "")
    channel_e9 = safe_get_cell(channel_platform, 8, 4, 0)
    channel_e10 = safe_get_cell(channel_platform, 9, 4, 0)
    channel_g5 = safe_get_cell(channel_platform, 4, 6, 0)
    channel_g6 = safe_get_cell(channel_platform, 5, 6, 0)
    channel_g7 = safe_get_cell(channel_platform, 6, 6, 0)
    channel_g8 = safe_get_cell(channel_platform, 7, 6, 0)
    channel_o5 = safe_get_cell(channel_platform, 4, 14, 0)
    channel_o6 = safe_get_cell(channel_platform, 5, 14, 0)
    channel_o7 = safe_get_cell(channel_platform, 6, 14, 0)
    channel_o8 = safe_get_cell(channel_platform, 7, 14, 0)
    channel_j9 = safe_get_cell(channel_platform, 8, 9, 0)
    channel_j10 = safe_get_cell(channel_platform, 9, 9, 0)
    channel_k9 = safe_get_cell(channel_platform, 8, 10, 0)
    channel_k10 = safe_get_cell(channel_platform, 9, 10, 0)
    channel_l9 = safe_get_cell(channel_platform, 8, 11, 0)
    channel_l10 = safe_get_cell(channel_platform, 9, 11, 0)

    program_l11 = safe_get_cell(program_performance, 10, 11, 0)
    program_l12 = safe_get_cell(program_performance, 11, 11, 0)
    program_f11 = safe_get_cell(program_performance, 10, 5, 0)
    program_g11 = safe_get_cell(program_performance, 10, 6, 0)
    program_f12 = safe_get_cell(program_performance, 11, 5, 0)
    program_g12 = safe_get_cell(program_performance, 11, 6, 0)

    current_year = datetime.now().year

    safe_set_cell(sheet2, 'B2', f"{prop_b1} - {current_year} Driven By: {prop_b29}")
    safe_set_cell(sheet2, 'C5', prop_b3)

    if isinstance(prop_b8, str):
        prop_b8 = datetime.strptime(prop_b8, "%d %B %Y")
    campaign_end_date = prop_b8 + timedelta(weeks=prop_b14)
    start_month = prop_b8.strftime("%b'%y")
    end_month = campaign_end_date.strftime("%b'%y")
    result = f"{start_month} - {end_month}"

        # ✅ 1. Load ER and CPRP Channels TV and Digital CTV-Mobile CPM.xlsx
    er_file_path = "input/ER and CPRP Channels TV and Digital CTV-Mobile CPM.xlsx"
    if not os.path.exists(er_file_path):
        logger.error(f"ER and CPRP Channels TV and Digital CTV-Mobile CPM.xlsx file not found at {er_file_path}")
        raise FileNotFoundError(f"ER and CPRP Channels TV and Digital CTV-Mobile CPM.xlsx file not found at {er_file_path}")

    er_dfa = pd.read_excel(er_file_path, sheet_name="ER Channels")
    er_dfb = pd.read_excel(er_file_path, sheet_name="CPRP Channels")

    # ✅ 2. Make sure columns exist
    if 'Channels' not in er_dfa.columns or 'Net Rate' not in er_dfa.columns:
        logger.error("ER and CPRP Channels TV and Digital CTV-Mobile CPM.xlsx must have 'Channels' and 'ER' columns.")
        raise ValueError("Missing required columns in ER and CPRP Channels TV and Digital CTV-Mobile CPM.xlsx")

    # ✅ 3. Get channel name from C6
    er_channel_name = channel_c6.strip().lower()

    # ✅ 4. Match & get ER value
    er_match = er_dfa[er_dfa['Channels'].str.strip().str.lower() == er_channel_name]

    if not er_match.empty:
        er_value = er_match.iloc[0]['Net Rate']
        logger.info(f"✅ ER value found for channel '{channel_c6}': {er_value}")
    else:
        er_value = "(ER not found)"
        logger.warning(f"⚠️ ER value NOT found for channel '{channel_c6}' in ER and CPRP Channels TV and Digital CTV-Mobile CPM.xlsx.")

    # ✅ 5. Write ER to M31 in 'DBD One Pager-with Eval.'
    safe_set_cell(sheet2, 'M31', er_value)
    logger.info(f"✅ ER value written to M31: {er_value}")
    
    if 'Channels' not in er_dfb.columns:
        logger.error("ER and CPRP Channels TV and Digital CTV-Mobile CPM.xlsx must have 'Channels' and 'ER' columns.")
        raise ValueError("Missing required columns in ER and CPRP Channels TV and Digital CTV-Mobile CPM.xlsx")
    
    er_channel_name2 = channel_c5.strip().lower()
    
    er_match2 = er_dfb[er_dfb['Channels'].str.strip().str.lower() == er_channel_name2]

    if not er_match2.empty:
        er_values = er_match2.iloc[0]['Market CPRP']
        logger.info(f"✅ ER value found for channel '{channel_c5}': {er_values}")
    else:
        er_values = "(ER not found)"
        logger.warning(f"⚠️ ER value NOT found for channel '{channel_c5}' in ER and CPRP Channels TV and Digital CTV-Mobile CPM.xlsx.")

    # ✅ 5. Write ER to M28 in 'DBD One Pager-with Eval.'
    safe_set_cell(sheet2, 'M28', er_values)
    logger.info(f"✅ ER value written to M28: {er_values}")

        # ✅ 6. Make sure 'All India CPRP' column exists in CPRP Channels sheet
    if 'All India CPRP' not in er_dfb.columns:
        logger.error("CPRP Channels sheet must have an 'All India CPRP' column.")
        raise ValueError("Missing 'All India CPRP' column in CPRP Channels sheet")

    # ✅ 7. Get any value under 'All India CPRP' (assuming all rows same)
    all_india_cprp_value = er_dfb['All India CPRP'].dropna().iloc[0]  # Pick first non-NaN value
    logger.info(f"✅ Extracted All India CPRP value: {all_india_cprp_value}")

    # ✅ 8. Write this CPRP value from N28 to N31
    for row in range(28, 32):
        cell_ref = f'N{row}'
        safe_set_cell(sheet2, cell_ref, all_india_cprp_value)
        logger.info(f"✅ Written All India CPRP value to {cell_ref}: {all_india_cprp_value}")



    # First, parse the dates from dd-mm-yyyy format
    start_dates = datetime.strptime(str(program_f12).strip(), "%Y-%m-%d %H:%M:%S")
    end_dates = datetime.strptime(str(program_g12).strip(), "%Y-%m-%d %H:%M:%S")

    # Format to Mon'YY
    start_formatted = start_dates.strftime("%b'%y")
    end_formatted = end_dates.strftime("%b'%y")

    # Combine
    formatted_range = f"{start_formatted} - {end_formatted}"

    # ✅ Write to both H21 and H22 (or wherever needed)
    safe_set_cell(sheet2, 'H21', formatted_range)
    safe_set_cell(sheet2, 'H22', formatted_range)

     # Row 10
    safe_set_cell(sheet2, 'D15', channel_c9) # C9
    safe_set_cell(sheet2, 'C10', result)
    safe_set_cell(sheet2, 'D10', prop_b11)  # B11
    safe_set_cell(sheet2, 'E10', prop_b7)   # B7
    safe_set_cell(sheet2, 'F10', f"{prop_b9} - {prop_b10}")  # B9 - B10
 
 # Row 21
    safe_set_cell(sheet2, 'C21', channel_c5)  # C5
    safe_set_cell(sheet2, 'C22', channel_c6)  # C6
    safe_set_cell(sheet2, 'D21', prop_b1)  # B1
    safe_set_cell(sheet2, 'D22', prop_b1)  # B1
    safe_set_cell(sheet2, 'E21', channel_c9)  # C9
    safe_set_cell(sheet2, 'E22', channel_c9)  # C9
    safe_set_cell(sheet2, 'G21', program_l11)  # L11
    safe_set_cell(sheet2, 'G22', program_l12)  # L12 

 # Rows 29-32 (Channel & Platform section)
    safe_set_cell(sheet2, 'C28', channel_c5)  # C5
    safe_set_cell(sheet2, 'C29', channel_c6)  # C6
    safe_set_cell(sheet2, 'C30', channel_c5)  # C7
    safe_set_cell(sheet2, 'C31', channel_c6)  # C8
 
    safe_set_cell(sheet2, 'D28', prop_b1)  # B1
    safe_set_cell(sheet2, 'D29', prop_b1)  # B1
    safe_set_cell(sheet2, 'D30', prop_b1)  # B1
    safe_set_cell(sheet2, 'D31', prop_b1)  # B1
 
    safe_set_cell(sheet2, 'E28', prop_b12 - 2)  # B12 - 2
    safe_set_cell(sheet2, 'E29', prop_b12 - 2)  # B12 - 2
    safe_set_cell(sheet2, 'E30', prop_b13)  # B13
    safe_set_cell(sheet2, 'E31', prop_b13)  # B13
 
    safe_set_cell(sheet2, 'F28', channel_o5)  # O5
    safe_set_cell(sheet2, 'F29', channel_o6)  # O6
    safe_set_cell(sheet2, 'F30', channel_o7)  # O7
    safe_set_cell(sheet2, 'F31', channel_o8)  # O8
 
 # Set formulas for calculated cells
    safe_set_cell(sheet2, 'G28', "=F28*E28")
    safe_set_cell(sheet2, 'G29', "=F29*E29")
    safe_set_cell(sheet2, 'G30', "=F30*E30")
    safe_set_cell(sheet2, 'G31', "=F31*E31")
    safe_set_cell(sheet2, 'G32', "=SUM(G28:G31)")

    safe_set_cell(sheet2, 'J28', "=I28*G28/10")
    safe_set_cell(sheet2, 'J29', "=I29*G29/10")
    safe_set_cell(sheet2, 'J30', "=I30*G30/10")
    safe_set_cell(sheet2, 'J31', "=I31*G31/10")
    safe_set_cell(sheet2, 'J32', "=SUM(J28:J31)")

 # For cells K29 to K32 (grouped)
    formula_k = "=L32/G32*10"
    safe_set_cell(sheet2, 'K28', formula_k)
    safe_set_cell(sheet2, 'K29', formula_k)
    safe_set_cell(sheet2, 'K30', formula_k)
    safe_set_cell(sheet2, 'K31', formula_k)
 
 # For cells L29 to L32 (grouped)
    formula_l = f"={prop_b32}*10000000"
    safe_set_cell(sheet2, 'L28', formula_l)
    safe_set_cell(sheet2, 'L29', formula_l)
    safe_set_cell(sheet2, 'L30', formula_l)
    safe_set_cell(sheet2, 'L31', formula_l)
    safe_set_cell(sheet2, 'L32', "=SUM(L28:L31)")
 
 # For cells O29 to O32 (data unavailable)
    safe_set_cell(sheet2, 'O28', "=J28*N28")
    safe_set_cell(sheet2, 'O29', "=J29*N29")
    safe_set_cell(sheet2, 'O30', "=J30*N30")
    safe_set_cell(sheet2, 'O31', "=J31*N31")
    safe_set_cell(sheet2, 'O32', "=SUM(O28:O31)")
 
 # Rows 38-39 (Second section)
    safe_set_cell(sheet2, 'C37', channel_c9)  # C9
    safe_set_cell(sheet2, 'C38', channel_c9)  # C9
 
    safe_set_cell(sheet2, 'D37', prop_b1)  # B1
    safe_set_cell(sheet2, 'D38', prop_b1)  # B1
 
    safe_set_cell(sheet2, 'E37', channel_c9)  # C9
    safe_set_cell(sheet2, 'E38', channel_c9)  # C9
 
    safe_set_cell(sheet2, 'F37', channel_e9)  # E9
    safe_set_cell(sheet2, 'F38', channel_e10)  # E10
 
    safe_set_cell(sheet2, 'G37', "=(I38*1000000)*0.6")  # 60% as decimal
    safe_set_cell(sheet2, 'G38', "=(I39*1000000)*0.6")
    safe_set_cell(sheet2, 'G39', "=SUM(G37:G38)")
 
    safe_set_cell(sheet2, 'H37', "=(I38*1000000)*0.4")  # 40% as decimal
    safe_set_cell(sheet2, 'H38', "=(I39*1000000)*0.4")
 
    safe_set_cell(sheet2, 'I37', channel_k9)  # K9
    safe_set_cell(sheet2, 'I38', channel_k10)  # K10
 
    safe_set_cell(sheet2, 'J37', channel_j9)  # J9
    safe_set_cell(sheet2, 'J38', channel_j10)  # J10
 
    safe_set_cell(sheet2, 'K37', channel_l9)  # L9
    safe_set_cell(sheet2, 'K38', channel_l10)  # L10
 
    safe_set_cell(sheet2, 'L37', "=(K37*I37/1000)*10^6")
    safe_set_cell(sheet2, 'L38', "=(K38*I38/1000)*10^6")
    safe_set_cell(sheet2, 'L39', "=SUM(L37:L38)")
 
 # For cells O38 to O39 (data unavailable)
    safe_set_cell(sheet2, 'O38', "(data unavailable)")
    safe_set_cell(sheet2, 'O39', "(data unavailable)")
 
 # Row 46
    #safe_set_cell(sheet2, 'D46', "=L33/10^7")
 
    #safe_set_cell(sheet2, 'E46', "=(L33+L40)/10^7")
 
 # Now fill data in Sheet 1 (Summary)
    logger.info("Filling Sheet 1: Summary")
 
 # B2 to K2 (grouped) = property details B1 - {current year} Driven By {property details B29}
    safe_set_cell(sheet1, 'B2', f"{prop_b1} - {current_year} Driven By {prop_b29}")
 
 # Various cells from Property Details
    safe_set_cell(sheet1, 'D4', result)  # property details B8
    safe_set_cell(sheet1, 'D5', prop_b1)  # property details B1
    safe_set_cell(sheet1, 'D6', prop_a29)  # property details A29
    safe_set_cell(sheet1, 'D7', prop_b4)  # property details B4
    safe_set_cell(sheet1, 'D10', prop_b20)  # property details B20
    safe_set_cell(sheet1, 'D11', prop_b23)  # property details B23
    safe_set_cell(sheet1, 'H10', prop_b21)  # property details B21
    safe_set_cell(sheet1, 'H11', prop_b22)  # property details B22
    safe_set_cell(sheet1, 'D14', prop_b26)  # property details B26
    safe_set_cell(sheet1, 'D15', prop_b29)  # property details B29
    safe_set_cell(sheet1, 'H14', prop_b27)  # property details B27
    safe_set_cell(sheet1, 'H15', prop_b28)  # property details B28
    safe_set_cell(sheet1, 'D19', f"{prop_b9} - {prop_b10} (Timing)")  # property details B9 - B10(timing)
 
 # Static text entries
    safe_set_cell(sheet1, 'D20', "TV Telecast - On")
    safe_set_cell(sheet1, 'D21', "Digital Telecast - On")
 
 # Channel & Platform details
    safe_set_cell(sheet1, 'C26', channel_c5)  # channel & platform details C5
    safe_set_cell(sheet1, 'C27', channel_c6)  # channel & platform details C6
    safe_set_cell(sheet1, 'C28', channel_c5)  # channel & platform details C7
    safe_set_cell(sheet1, 'C29', channel_c6)  # channel & platform details C8
 
 # Property details for rows 26-29
    safe_set_cell(sheet1, 'D26', prop_b12 - 2)  # property details {value in (B12) -2}
    safe_set_cell(sheet1, 'D27', prop_b12 - 2)  # property details {value in (B12) -2}
    safe_set_cell(sheet1, 'D28', prop_b13)  # property details B13
    safe_set_cell(sheet1, 'D29', prop_b13)  # property details B13
 
    safe_set_cell(sheet1, 'E26', prop_b1)  # property details B1
    safe_set_cell(sheet1, 'E27', prop_b1)  # property details B1
    safe_set_cell(sheet1, 'E28', prop_b1)  # property details B1
    safe_set_cell(sheet1, 'E29', prop_b1)  # property details B1
 
 # E46 = =(L33+L40)/10^7
 
    safe_set_cell(sheet1, 'F26', channel_g5)  # channel_g5
    safe_set_cell(sheet1, 'F27', channel_g6)  # channel_g6
    safe_set_cell(sheet1, 'F28', channel_g7)  # channel_g7
    safe_set_cell(sheet1, 'F29', channel_g8)  # channel_g8

 # Channel & Platform details for rows 34-35
    safe_set_cell(sheet1, 'C34', channel_c9)  # channel & platform details C9
    safe_set_cell(sheet1, 'C35', channel_c9)  # channel & platform details C9
    safe_set_cell(sheet1, 'D34', channel_e9)  # channel & platform details E9
    safe_set_cell(sheet1, 'D35', channel_e10)  # channel & platform details E10
    safe_set_cell(sheet1, 'E34', prop_b1)  # property details B1
    safe_set_cell(sheet1, 'E35', prop_b1)  # property details B1
    safe_set_cell(sheet1, 'F34', channel_j9)  # channel & platform details J9
    safe_set_cell(sheet1, 'F35', channel_j10)  # channel & platform details J10
    safe_set_cell(sheet1, 'G34', channel_k9)  # channel & platform details K9
    safe_set_cell(sheet1, 'G26', channel_b5)  # channel & platform details B5
    safe_set_cell(sheet1, 'G27', channel_b5)  # channel & platform details B5
    safe_set_cell(sheet1, 'G28', channel_b5)  # channel & platform details B5
    safe_set_cell(sheet1, 'G29', channel_b5)  # channel & platform details B5
    safe_set_cell(sheet1, 'G35', channel_k10)  # channel & platform details K10
 
 # D41 to E41 (grouped) = Sheet1 D46
 # We'll copy the formula from Sheet2 D46
 #sheet2_d46_formula = "=L33/10^7"  # Same formula as in Sheet2
 #safe_set_cell(sheet1, 'D41', sheet2_d46_formula)
 
    safe_set_cell(sheet1, 'D41', "='DBD One Pager-with Eval.'!D45")
 
    safe_set_cell(sheet1, 'F41', "='DBD One Pager-with Eval.'!E45")

    # ✅✅✅ Call TVR Extractor and write to H30 and I30:
    tvrs = extract_tvr_data(input_a_path)
    if tvrs and len(tvrs) >= 4:
        safe_set_cell(sheet2, 'I28', tvrs[0])
        safe_set_cell(sheet2, 'I29', tvrs[1])
        safe_set_cell(sheet2, 'I30', tvrs[0])
        safe_set_cell(sheet2, 'I31', tvrs[1])
        safe_set_cell(sheet2, 'H28', tvrs[2])
        safe_set_cell(sheet2, 'H29', tvrs[3])
        safe_set_cell(sheet2, 'H30', tvrs[2])
        safe_set_cell(sheet2, 'H31', tvrs[3])
        logger.info(f"✅ TVRs written: I28={tvrs[0]}, I29={tvrs[1]}, I30={tvrs[0]}, I31={tvrs[1]}, H28={tvrs[2]}, H29={tvrs[3]}, H30={tvrs[2]}, H31={tvrs[3]}")
    else:
        logger.warning("⚠️ No TVRs returned to write in H30, I30.")

    # ✅ Save workbook
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    logger.info(f"✅ Process finished. Output saved to {output_path}")
    print(f"✅ Detailed Package file generated successfully")

# ✅✅✅ Final: main block — ye sabke last mei lagao

if __name__ == "__main__":
    non_cricket_folder = "input/non_cricket_input/"
    input_b = "input/TVR Output.xlsx"
    skeleton = "input/Skeleton Output.xlsx"
    output = "output/Completed_Output.xlsx"
    
    non_cricket_files = glob.glob(os.path.join(non_cricket_folder, "*.xlsx"))
    if not non_cricket_files:
        raise FileNotFoundError(f"No Non Cricket Input file found in {non_cricket_folder}")
    
    input_a = non_cricket_files[0]  # Pick first match

    print(f"Using Non Cricket Input file: {input_a}")

    process_excel_data(input_a, input_b, skeleton, output)
